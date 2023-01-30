import csv
import glob
from openpyxl import Workbook
from pathlib import Path

wb = Workbook()
ws = wb.active
col = 1
track = input("Введите к папке с файлами: ")
for filename in glob.glob(track + "\\*.txt"):
    rows = 3
    with open(filename) as f:
        f_csv = csv.reader(f, delimiter="\t")
        next(f)
        f_csv_head = next(f_csv)
        hzs, lev_db_s = [], []
        for row in f_csv:
            hz = row[0]
            lev_db = float(row[1])
            hzs.append(hz)
            lev_db_s.append(lev_db)
        if col == 1:
            ws.cell(row=2, column=col, value=f_csv_head[0])
            for hz in hzs:
                ws.cell(row=rows, column=col, value=hz)
                rows += 1
            rows = 3
            col += 1
        ws.cell(row=1, column=col, value=Path(filename).stem)
        ws.cell(row=2, column=col, value=f_csv_head[1])
        for db in lev_db_s:
            ws.cell(row=rows, column=col, value=db)
            rows += 1
        rows = 3
        col += 1

wb.save(track + '//Сведенные данные.xlsx')
